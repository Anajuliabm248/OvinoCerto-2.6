"""
Importa as exigências nutricionais NRC (2007) da planilha Excel.

Uso:
    python manage.py seed_exigencias
    python manage.py seed_exigencias --excel /caminho/para/arquivo.xlsx
    python manage.py seed_exigencias --limpar  # apaga tudo antes de importar
"""
import os
from django.core.management.base import BaseCommand, CommandError


# Mapeamento: nome da categoria no Excel 
CATEGORIA_MAP = {
    'Cordeiros(as) (4 meses)': 'cordeiros_4_meses',
    'Cordeiros(as) (8 meses)': 'cordeiros_8_meses',
    'Carneiros (4 meses)':     'carneiro_4_meses',
    'Carneiros (8 meses)':     'carneiro_8_meses',
    'Carneiros':               'carneiros',
    'Ovelhas':                 'ovelhas',
}

# Mapeamento: nome da fase no Excel
FASE_MAP = {
    'Crescimento':       'crescimento',
    'Manutenção':        'manutencao',
    'Pré-cobrição':      'pre_cobricao',
    'Reprodução':        'reproducao',
    'Gestação precoce':  'gestacao_precoce',
    'Gestação tardia':   'gestacao_tardia',
    'Início da lactação': 'inicio_lactacao',
    'Meio da lactação':  'meio_lactacao',
    'Lactação tardia':   'lactacao_tardia',
    # Variações com acentos diferentes
    'Gestaçao precoce':  'gestacao_precoce',
    'Gestaçao tardia':   'gestacao_tardia',
    'Inicio da lactaçao': 'inicio_lactacao',
    'Meio da lactaçao':  'meio_lactacao',
    'Lactaçao tardia':   'lactacao_tardia',
}

# Colunas do Excel (índices base-0):
# 0=N°, 1=Categoria, 2=Fase, 3=PV, 4=Tipo_Parto, 5=PV_Nascer, 6=GMD,
# 7=%PV, 8=CMS, 9=PB_g, 10=%PB, 11=NDT_kg, 12=%NDT, 13=FDN_kg, 14=FDN%,
# 15=EE_kg, 16=EE%, 17=Ca_g, 18=%Ca, 19=P_g, 20=%P, 21=Ca/P
COL_IDX = {
    'numero':      0,
    'categoria':   1,
    'fase':        2,
    'pv_kg':       3,
    'tipo_parto':  4,
    'pv_nascer':   5,
    'gmd_kg':      6,
    'pv_percent':  7,
    'cms_kg':      8,
    'pb_g':        9,
    'pb_percent':  10,
    'ndt_kg':      11,
    'ndt_percent': 12,
    'fdn_kg':      13,
    'fdn_percent': 14,
    'ee_kg':       15,
    'ee_percent':  16,
    'ca_g':        17,
    'ca_percent':  18,
    'p_g':         19,
    'p_percent':   20,
    'ca_p':        21,
}

EXCEL_PADRAO = os.path.join(
    os.path.dirname(__file__),              # commands/
    '..', '..', '..', '..', '..',           # sobe até raiz do projeto
    'OvinoCerto_CORDEIRO_BaseProgramaExcel__ARRUMADO0604.xlsx',
)


def _float(val):
    """Converte valor do Excel para float; retorna None se inválido."""
    if val is None:
        return None
    if isinstance(val, str) and val.strip().startswith('-'):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _tipo_parto(val):
    """Converte valor do Excel para inteiro de tipo de parto; retorna None se inválido."""
    if val is None:
        return None
    if isinstance(val, str) and val.strip().startswith('-'):
        return None
    try:
        v = int(float(val))
        return v if 1 <= v <= 5 else None
    except (TypeError, ValueError):
        return None


class Command(BaseCommand):
    help = 'Importa as exigências NRC do arquivo Excel para o banco de dados.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel',
            default=EXCEL_PADRAO,
            help='Caminho para o arquivo .xlsx (default: raiz do projeto)',
        )
        parser.add_argument(
            '--limpar',
            action='store_true',
            help='Apaga todos os registros existentes antes de importar.',
        )
        parser.add_argument(
            '--aba',
            default='Exigências Nutricionais',
            help='Nome da aba no Excel (default: "Exigências Nutricionais")',
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError('openpyxl não instalado. Execute: pip install openpyxl')

        excel_path = os.path.abspath(options['excel'])
        if not os.path.isfile(excel_path):
            raise CommandError(f'Arquivo não encontrado: {excel_path}')

        self.stdout.write(f'Lendo: {excel_path}')
        wb = load_workbook(excel_path, read_only=True, data_only=True)

        aba = options['aba']
        if aba not in wb.sheetnames:
            raise CommandError(
                f'Aba "{aba}" não encontrada. Abas disponíveis: {wb.sheetnames}'
            )
        ws = wb[aba]

        from exigencia_nrc.models import ExigenciaNRC

        if options['limpar']:
            deleted, _ = ExigenciaNRC.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'{deleted} registros removidos.'))

        criados = 0
        ignorados = 0

        for row in ws.iter_rows(min_row=5, values_only=True):
            # Ignora linhas sem número válido (cabeçalho, vazias, etc.)
            numero = row[COL_IDX['numero']]
            if not numero or not isinstance(numero, (int, float)):
                continue

            cat_excel  = str(row[COL_IDX['categoria']] or '').strip()
            fase_excel = str(row[COL_IDX['fase']] or '').strip()

            categoria = CATEGORIA_MAP.get(cat_excel)
            fase      = FASE_MAP.get(fase_excel)

            if not categoria:
                self.stdout.write(
                    self.style.WARNING(f'  Linha {numero}: categoria desconhecida "{cat_excel}" — ignorada.')
                )
                ignorados += 1
                continue

            if not fase:
                self.stdout.write(
                    self.style.WARNING(f'  Linha {numero}: fase desconhecida "{fase_excel}" — ignorada.')
                )
                ignorados += 1
                continue

            pv_kg = _float(row[COL_IDX['pv_kg']])
            if pv_kg is None:
                ignorados += 1
                continue

            obj = ExigenciaNRC(
                categoria            = categoria,
                fase                 = fase,
                pv_kg                = pv_kg,
                tipo_parto           = _tipo_parto(row[COL_IDX['tipo_parto']]),
                pv_nascer_kg         = _float(row[COL_IDX['pv_nascer']]),
                gmd_kg               = _float(row[COL_IDX['gmd_kg']]),
                pv_percentual        = _float(row[COL_IDX['pv_percent']]),
                cms_kg               = _float(row[COL_IDX['cms_kg']]),
                pb_g                 = _float(row[COL_IDX['pb_g']]),
                pb_percentual        = _float(row[COL_IDX['pb_percent']]),
                ndt_kg               = _float(row[COL_IDX['ndt_kg']]),
                ndt_percentual       = _float(row[COL_IDX['ndt_percent']]),
                fdn_kg               = _float(row[COL_IDX['fdn_kg']]),
                fdn_percentual       = _float(row[COL_IDX['fdn_percent']]),
                ee_kg                = _float(row[COL_IDX['ee_kg']]),
                ee_percentual        = _float(row[COL_IDX['ee_percent']]),
                ca_g                 = _float(row[COL_IDX['ca_g']]),
                ca_percentual        = _float(row[COL_IDX['ca_percent']]),
                p_g                  = _float(row[COL_IDX['p_g']]),
                p_percentual         = _float(row[COL_IDX['p_percent']]),
                ca_p_percentual      = _float(row[COL_IDX['ca_p']]),
            )
            obj.save()
            criados += 1

        self.stdout.write(
            self.style.SUCCESS(
                f'\nConcluído: {criados} exigências importadas, {ignorados} ignoradas.'
            )
        )