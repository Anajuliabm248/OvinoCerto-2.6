"""
Importa as exigências nutricionais NRC (2007) da planilha Excel.

Uso:
    python manage.py seed_exigencias
    python manage.py seed_exigencias --excel backend/base_ovino.xls
    python manage.py seed_exigencias --limpar   # apaga todas as exigências antes de importar
"""
import os
import unicodedata
from django.core.management.base import BaseCommand, CommandError

# pylint: skip-file

# Mapeamento: nome da categoria no Excel
CATEGORIA_MAP = {
    'Cordeiros(as) (4 meses)': 'cordeiros_4_meses',
    'Cordeiros(as) (8 meses)': 'cordeiros_8_meses',
    'Carneiros (4 meses)':     'carneiro_4_meses',
    'Carneiros (8 meses)':     'carneiro_8_meses',
    'Carneiros':               'carneiros',
    'Ovelhas':                 'ovelhas',
}

# Mapeamento: nome da fase no Excel
FASE_MAP = {
    'Crescimento':       'crescimento',
    'Manutenção':        'manutencao',
    'Pré-cobrição':      'pre_cobricao',
    'Reprodução':        'reproducao',
    'Gestação precoce':  'gestacao_precoce',
    'Gestação tardia':   'gestacao_tardia',
    'Início de lactação': 'inicio_lactacao',
    'Meio de lactação':  'meio_lactacao',
    'Lactação tardia':   'lactacao_tardia',
    # Variações com acentos diferentes
    'Gestaçao precoce':  'gestacao_precoce',
    'Gestaçao tardia':   'gestacao_tardia',
    'Inicio de lactaçao': 'inicio_lactacao',
    'Meio de lactaçao':  'meio_lactacao',
    'Lactaçao tardia':   'lactacao_tardia',
}

TIPO_PARTO_MAP = {
    '1 Cordeiro':  1,
    '2 Cordeiro':  2,
    '2 Cordeiros': 2,
    '3 Cordeiro':  3,
    '3 Cordeiros': 3,
    '4 Cordeiro':  4,
    '4 Cordeiros': 4,
    '5 Cordeiro':  5,
    '5 Cordeiros': 5,
}

FASES_GESTACAO = {'gestacao_precoce', 'gestacao_tardia'}
FASES_LACTACAO = {'inicio_lactacao', 'meio_lactacao', 'lactacao_tardia'}


def _normalize(s: str) -> str:
    """Normalize string for comparison: lowercase, strip, remove accents, collapse spaces."""
    if s is None:
        return ''
    if not isinstance(s, str):
        s = str(s)
    s = s.strip().lower()
    # Remove accents
    s = unicodedata.normalize('NFD', s)
    s = ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')
    # Collapse whitespace
    s = ' '.join(s.split())
    return s

# Precompute normalized maps for resilient matching
NORM_CATEGORIA_MAP = { _normalize(k): v for k, v in CATEGORIA_MAP.items() }
NORM_FASE_MAP = { _normalize(k): v for k, v in FASE_MAP.items() }
NORM_TIPO_PARTO_MAP = { _normalize(k): v for k, v in TIPO_PARTO_MAP.items() }


def _tipo_parto(val):
    return NORM_TIPO_PARTO_MAP.get(_normalize(val))

# Colunas do Excel (índices base-0):
COL_IDX = {
    'numero':      0,
    'categoria':   1,
    'fase':        2,
    'pv_kg':       3,
    'tipo_parto':  4,
    'pv_nascer_ou_leite': 5,
    'gmd_kg':      6,
    'pv_percent':  7,
    'cms_kg':      8,
    'pb_g':        9,
    'pb_percent':  10,
    'ndt_kg':      11,
    'ndt_percent': 12,
    'fdn_kg':      13,
    'fdn_percent': 14,
    'ee_kg':       15,
    'ee_percent':  16,
    'ca_g':        17,
    'ca_percent':  18,
    'p_g':         19,
    'p_percent':   20,
    'ca_p':        21,
}

EXCEL_PADRAO = os.path.join(
    os.path.dirname(__file__),
    '..',
    '..',
    '..',
    'base_ovino.xls',
)


def _float(val):
    """Converte valor do Excel para float; retorna None se inválido."""
    if val is None:
        return None
    if isinstance(val, str) and val.strip().startswith('-'):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


class Command(BaseCommand):
    help = 'Importa as exigências NRC do arquivo Excel para o banco de dados.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel',
            default=EXCEL_PADRAO,
            help='Caminho para o arquivo .xlsx (default: raiz do projeto)',
        )
        parser.add_argument(
            '--limpar',
            action='store_true',
            help='Apaga todos os registros existentes antes de importar.',
        )
        parser.add_argument(
            '--aba',
            default='Exigências Nutricionais',
            help='Nome da aba no Excel (default: "Exigências Nutricionais")',
        )

    def handle(self, *args, **options):
        excel_path = os.path.abspath(options['excel'])
        if not os.path.isfile(excel_path):
            raise CommandError(f'Arquivo não encontrado: {excel_path}')

        self.stdout.write(f'Lendo: {excel_path}')
        _, ext = os.path.splitext(excel_path)
        ext = ext.lower()

        if ext == '.xls':
            try:
                import xlrd
            except ImportError:
                raise CommandError('xlrd não instalado. Execute: pip install xlrd')

            wb_xl = xlrd.open_workbook(excel_path, formatting_info=False)
            sheet_names = wb_xl.sheet_names()
            aba = options['aba']
            if aba not in sheet_names:
                match = next((s for s in sheet_names if 'Exig' in s or 'Nutri' in s), None)
                if match:
                    aba = match
                    self.stdout.write(self.style.WARNING(f'Aba não encontrada; usando "{aba}"'))
                else:
                    raise CommandError(
                        f'Aba "{aba}" não encontrada. Abas disponíveis: {sheet_names}'
                    )
            ws = wb_xl.sheet_by_name(aba)

            def rows_iterator():
                # original used min_row=5 (1-based), so start at index 4
                for r in range(4, ws.nrows):
                    yield tuple(ws.cell_value(r, c) for c in range(ws.ncols))

            rows = rows_iterator()
        else:
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise CommandError('openpyxl não instalado. Execute: pip install openpyxl')

            wb = load_workbook(excel_path, read_only=True, data_only=True)

            aba = options['aba']
            if aba not in wb.sheetnames:
                match = next((s for s in wb.sheetnames if 'Exig' in s or 'Nutri' in s), None)
                if match:
                    aba = match
                    self.stdout.write(self.style.WARNING(f'Aba não encontrada; usando "{aba}"'))
                else:
                    raise CommandError(
                        f'Aba "{aba}" não encontrada. Abas disponíveis: {wb.sheetnames}'
                    )
            ws = wb[aba]
            rows = ws.iter_rows(min_row=5, values_only=True)

        from exigencia_nrc.models import ExigenciaNRC

        if options['limpar']:
            deleted, _ = ExigenciaNRC.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'{deleted} registros removidos.'))

        criados = 0
        ignorados = 0

        for row in rows:
            # Ignora linhas sem número válido (cabeçalho, vazias, etc.)
            numero = row[COL_IDX['numero']]
            if not numero or not isinstance(numero, (int, float)):
                continue

            cat_excel  = row[COL_IDX['categoria']]
            fase_excel = row[COL_IDX['fase']]

            cat_norm = _normalize(cat_excel)
            fase_norm = _normalize(fase_excel)

            categoria = NORM_CATEGORIA_MAP.get(cat_norm)
            fase      = NORM_FASE_MAP.get(fase_norm)

            if not categoria:
                self.stdout.write(
                    self.style.WARNING(f'  Linha {numero}: categoria desconhecida "{cat_excel}" — ignorada.')
                )
                ignorados += 1
                continue

            if not fase:
                self.stdout.write(
                    self.style.WARNING(f'  Linha {numero}: fase desconhecida "{fase_excel}" — ignorada.')
                )
                ignorados += 1
                continue

            pv_kg = _float(row[COL_IDX['pv_kg']])
            if pv_kg is None:
                ignorados += 1
                continue

            valor_col_compartilhada = _float(row[COL_IDX['pv_nascer_ou_leite']])
            pv_nascer_kg = valor_col_compartilhada if fase in FASES_GESTACAO else None
            producao_leite_kg_dia = valor_col_compartilhada if fase in FASES_LACTACAO else None

            obj = ExigenciaNRC(
                categoria            = categoria,
                fase                 = fase,
                pv_kg                = pv_kg,
                tipo_parto           = _tipo_parto(row[COL_IDX['tipo_parto']]),
                pv_nascer_kg         = pv_nascer_kg,
                producao_leite_kg_dia = producao_leite_kg_dia,
                gmd_kg               = _float(row[COL_IDX['gmd_kg']]),
                pv_percentual        = _float(row[COL_IDX['pv_percent']]),
                cms_kg               = _float(row[COL_IDX['cms_kg']]),
                pb_g                 = _float(row[COL_IDX['pb_g']]),
                pb_percentual        = _float(row[COL_IDX['pb_percent']]),
                ndt_kg               = _float(row[COL_IDX['ndt_kg']]),
                ndt_percentual       = _float(row[COL_IDX['ndt_percent']]),
                fdn_kg               = _float(row[COL_IDX['fdn_kg']]),
                fdn_percentual       = _float(row[COL_IDX['fdn_percent']]),
                ee_kg                = _float(row[COL_IDX['ee_kg']]),
                ee_percentual        = _float(row[COL_IDX['ee_percent']]),
                ca_g                 = _float(row[COL_IDX['ca_g']]),
                ca_percentual        = _float(row[COL_IDX['ca_percent']]),
                p_g                  = _float(row[COL_IDX['p_g']]),
                p_percentual         = _float(row[COL_IDX['p_percent']]),
                ca_p_percentual      = _float(row[COL_IDX['ca_p']]),
            )
            obj.save()
            criados += 1

        self.stdout.write(
            self.style.SUCCESS(
                f'\nConcluído: {criados} exigências importadas, {ignorados} ignoradas.'
            )
        )
