"""
Importa os ingredientes da tabela Valadares Filho (2010) do Excel.

Uso:
    python manage.py seed_ingredientes
    python manage.py seed_ingredientes --excel /caminho/para/arquivo.xlsx
    python manage.py seed_ingredientes --limpar   # apaga Valadares antes de importar
"""
import os
from django.core.management.base import BaseCommand, CommandError


# Mapeamento: classificação no Excel → chave Django
CLASSIF_MAP = {
    'Volumoso':    'volumoso',
    'Concentrado': 'concentrado',
}

# Mapeamento: tipo (subtipo) no Excel → chave Django
TIPO_MAP = {
    'Forragens Secas':  'forragens_secas',
    'Forragens Verdes': 'forragens_verdes',
    'Silagens':         'silagens',
    'Energético':       'energetico',
    'Proteico':         'proteico',
    'Mineral':          'mineral',
    'Aditivos':         'aditivos',
}

# Colunas do Excel (índices base-0):
# 0=N°, 1=Classificação, 2=Tipo, 3=Ingrediente, 4=MS%, 5=PB%, 6=NDT%, 7=FDN%, 8=EE%, 9=Ca%, 10=P%, 11=R$/kg
COL_IDX = {
    'numero':       0,
    'classificacao':1,
    'tipo':         2,
    'nome':         3,
    'ms':           4,
    'pb':           5,
    'ndt':          6,
    'fdn':          7,
    'ee':           8,
    'ca':           9,
    'p':            10,
    'custo_kg':     11,
}

EXCEL_PADRAO = os.path.join(
    os.path.dirname(__file__),
    '..', '..', '..', '..', '..',
    'OvinoCerto_CORDEIRO_BaseProgramaExcel__ARRUMADO0604.xlsx',
)


def _float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


class Command(BaseCommand):
    help = 'Importa ingredientes da tabela Valadares Filho (2010) do arquivo Excel.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel',
            default=EXCEL_PADRAO,
            help='Caminho para o arquivo .xlsx',
        )
        parser.add_argument(
            '--limpar',
            action='store_true',
            help='Remove todos os ingredientes Valadares antes de importar.',
        )
        parser.add_argument(
            '--aba',
            default='Composição Bromatológica e Cust',
            help='Nome da aba no Excel',
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError('openpyxl não instalado. Execute: pip install openpyxl')

        excel_path = os.path.abspath(options['excel'])
        if not os.path.isfile(excel_path):
            raise CommandError(f'Arquivo não encontrado: {excel_path}')

        self.stdout.write(f'Lendo: {excel_path}')
        wb = load_workbook(excel_path, read_only=True, data_only=True)

        aba = options['aba']
        if aba not in wb.sheetnames:
            # Tenta encontrar aba com nome similar
            match = next((s for s in wb.sheetnames if 'Bromatol' in s or 'Cust' in s), None)
            if match:
                aba = match
                self.stdout.write(self.style.WARNING(f'Aba não encontrada; usando "{aba}"'))
            else:
                raise CommandError(
                    f'Aba "{aba}" não encontrada. Abas disponíveis: {wb.sheetnames}'
                )
        ws = wb[aba]

        from ingrediente.models import Ingrediente

        if options['limpar']:
            deleted, _ = Ingrediente.objects.filter(fonte_valadares=True).delete()
            self.stdout.write(self.style.WARNING(f'{deleted} ingredientes Valadares removidos.'))

        criados = 0
        atualizados = 0
        ignorados = 0
        tipo_atual = None  # O Excel agrupa por tipo em linhas mescladas

        for row in ws.iter_rows(min_row=3, values_only=True):
            numero = row[COL_IDX['numero']]
            if not numero or not isinstance(numero, (int, float)):
                # Pode ser uma linha de subtítulo de tipo (ex: "Forragens Verdes")
                # Captura para usar nas próximas linhas
                tipo_cell = row[COL_IDX['tipo']]
                if tipo_cell and isinstance(tipo_cell, str) and tipo_cell.strip() in TIPO_MAP:
                    tipo_atual = tipo_cell.strip()
                continue

            classif_excel = str(row[COL_IDX['classificacao']] or '').strip()
            tipo_excel    = str(row[COL_IDX['tipo']]          or '').strip()
            nome          = str(row[COL_IDX['nome']]          or '').strip()

            if not nome:
                ignorados += 1
                continue

            classificacao = CLASSIF_MAP.get(classif_excel)
            if not classificacao:
                # Herda da última linha válida de classificação
                if classif_excel in ('', None) and tipo_atual:
                    # Tenta inferir classificação pelo tipo
                    from ingrediente.models import TIPO_CHOICES
                    tipo_key = TIPO_MAP.get(tipo_excel or tipo_atual)
                    if tipo_key in ('forragens_secas', 'forragens_verdes', 'silagens'):
                        classificacao = 'volumoso'
                    else:
                        classificacao = 'concentrado'
                else:
                    self.stdout.write(
                        self.style.WARNING(
                            f'  Linha {numero}: classificação desconhecida "{classif_excel}" — ignorada.'
                        )
                    )
                    ignorados += 1
                    continue

            tipo_key = TIPO_MAP.get(tipo_excel) or TIPO_MAP.get(tipo_atual)
            if not tipo_key:
                tipo_key = 'outro'

            defaults = {
                'classificacao':  classificacao,
                'tipo':           tipo_key,
                'ms':             _float(row[COL_IDX['ms']]),
                'pb':             _float(row[COL_IDX['pb']]),
                'ndt':            _float(row[COL_IDX['ndt']]),
                'fdn':            _float(row[COL_IDX['fdn']]),
                'ee':             _float(row[COL_IDX['ee']]),
                'ca':             _float(row[COL_IDX['ca']]),
                'p':              _float(row[COL_IDX['p']]),
                'custo_kg':       _float(row[COL_IDX['custo_kg']]),
                'fonte_valadares': True,
                'usuario':        None,
            }

            obj, created = Ingrediente.objects.update_or_create(
                nome=nome,
                fonte_valadares=True,
                defaults=defaults,
            )

            if created:
                criados += 1
            else:
                atualizados += 1

            # Atualiza tipo_atual para linhas seguintes sem subtítulo explícito
            if tipo_excel:
                tipo_atual = tipo_excel

        self.stdout.write(
            self.style.SUCCESS(
                f'\nConcluído: {criados} criados, {atualizados} atualizados, {ignorados} ignorados.'
            )
        )